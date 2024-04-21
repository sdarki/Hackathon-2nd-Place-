import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time

def get_news_titles(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        titles = soup.find_all(class_='WavNE')  
        return [title.get_text() for title in titles]
    else:
        print("Failed to fetch content from the website.")
        return []


def write_to_excel(titles):
    try:
        wb = load_workbook("news_data.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Text", "Label"]) 
    
    ws = wb.active
    existing_titles = {cell.value for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row) for cell in row}  
    for title in titles:
        if title not in existing_titles:
            ws.append([title, title, 1])  
            existing_titles.add(title)  
    
    wb.save("news_data.xlsx")
    print("New unique titles have been appended to news_data.xlsx")

# Main function
def main():
    url = "https://timesofindia.indiatimes.com/india"  
    previous_titles = set()
    
    while True:
        current_titles = set(get_news_titles(url))
        new_titles = current_titles - previous_titles
        
        if new_titles:
            write_to_excel(new_titles)
            previous_titles = current_titles
        
        time.sleep(3600) 

if _name_ == "_main_":
    main()
